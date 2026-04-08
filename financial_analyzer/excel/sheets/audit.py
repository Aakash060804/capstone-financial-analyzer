"""
Audit trail sheet — logs run metadata and anomaly detection report.
"""

import datetime
from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from excel.styles       import (
    style_header, thin_border, set_col_width,
    set_row_height, hide_gridlines, NAVY, ALT_ROW,
)
from config.settings import (
    COMPANY_NAME, SCREENER_SLUG, SCREENER_URL,
    LLM_MODEL, FORECAST_YEARS, DCF_WACC,
    DCF_TERMINAL_GROWTH, SCENARIO_ASSUMPTIONS,
)
from utils.logger import get_logger

logger = get_logger(__name__)

# Severity colors
SEV_COLORS = {
    "ERROR":   {"text": "8B0000", "bg": "FDECEA"},
    "WARNING": {"text": "C55A11", "bg": "FFF3E0"},
    "INFO":    {"text": "1F4E79", "bg": "EBF5FB"},
}


def build_audit_sheet(
    wb: Workbook,
    canon_shape: tuple,
    ratio_shape: tuple,
    anomaly_report: dict = None,
    wacc_result:       dict = None,
    clustering_result: dict = None,
):
    ws = wb.create_sheet("Audit Trail")
    hide_gridlines(ws)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("B2:H2")
    style_header(ws.cell(2, 2), "Audit Trail  |  Run Metadata", size=11)
    set_row_height(ws, 2, 22)

    # ── Run metadata entries ───────────────────────────────────────────────────
    entries = [
        ("Run Timestamp",       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Company Name",        COMPANY_NAME),
        ("Screener Slug",       SCREENER_SLUG),
        ("Data Source URL",     SCREENER_URL),
        ("LLM Model",           LLM_MODEL),
        ("Canonical Metrics",   f"{canon_shape[0]} metrics × {canon_shape[1]} years"),
        ("Ratios Computed",     f"{ratio_shape[0]} ratios × {ratio_shape[1]} years"),
        ("Forecast Years",      str(FORECAST_YEARS)),
        ("DCF WACC",            f"{wacc_result['wacc']*100:.2f}%  (dynamic CAPM)" if wacc_result else f"{DCF_WACC*100:.1f}%  (fixed)"),
        ("DCF Terminal Growth", f"{DCF_TERMINAL_GROWTH*100:.1f}%"),
        ("Base Revenue Growth", f"{SCENARIO_ASSUMPTIONS['base']['revenue_growth']*100:.1f}%"),
        ("Bull Revenue Growth", f"{SCENARIO_ASSUMPTIONS['bull']['revenue_growth']*100:.1f}%"),
        ("Bear Revenue Growth", f"{SCENARIO_ASSUMPTIONS['bear']['revenue_growth']*100:.1f}%"),
    ]

    for ri, (label, value) in enumerate(entries):
        row_num = 4 + ri
        alt     = (ri % 2 == 0)
        fill    = PatternFill("solid", fgColor=ALT_ROW) if alt else None

        lc = ws.cell(row_num, 2)
        vc = ws.cell(row_num, 5)

        lc.value     = f"  {label}"
        lc.font      = Font(name="Arial", size=9, bold=True)
        lc.border    = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"B{row_num}:D{row_num}")
        if fill:
            lc.fill = fill

        vc.value     = value
        vc.font      = Font(name="Arial", size=9)
        vc.border    = thin_border()
        vc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"E{row_num}:H{row_num}")
        if fill:
            vc.fill = fill

        set_row_height(ws, row_num, 15)

    # ── DATA QUALITY REPORT ───────────────────────────────────────────────────
    if anomaly_report:
        _build_data_quality_section(ws, anomaly_report, start_row=4 + len(entries) + 2)

    # ── INDUSTRY CLASSIFICATION SECTION ───────────────────────────────────────
    if clustering_result:
        _build_clustering_section(ws, clustering_result,
            start_row=4 + len(entries) + 2 + _anomaly_rows(anomaly_report))

    # ── WACC COMPUTATION SECTION ──────────────────────────────────────────────
    if wacc_result:
        _build_wacc_section(ws, wacc_result,
            start_row=4 + len(entries) + 2 + _anomaly_rows(anomaly_report) + _clustering_rows(clustering_result))

    # ── Column widths ──────────────────────────────────────────────────────────
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 8)
    for ci in range(3, 9):
        set_col_width(ws, ci, 15)

    logger.info("Sheet 'Audit Trail' written")


def _build_data_quality_section(ws, anomaly_report: dict, start_row: int):
    """
    Appends the DATA QUALITY REPORT section to the Audit Trail sheet.
    """
    summary = anomaly_report.get("summary", {})
    flags   = anomaly_report.get("flags", [])
    row     = start_row

    # ── Section header ─────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:N{row}")
    h = ws.cell(row, 2)
    h.value     = "DATA QUALITY REPORT  |  AI-Powered Anomaly Detection  |  Z-Score + Isolation Forest"
    h.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor=NAVY)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # ── Summary row ────────────────────────────────────────────────────────────
    score          = summary.get("integrity_score", 100.0)
    score_color    = "1E6B1E" if score >= 95 else ("C55A11" if score >= 80 else "8B0000")
    score_bg       = "E8F5E9" if score >= 95 else ("FFF3E0" if score >= 80 else "FDECEA")

    summary_text = (
        f"  Metrics scanned: {summary.get('metrics_scanned', 0)}  |  "
        f"Years scanned: {summary.get('years_scanned', 0)}  |  "
        f"Total flags: {summary.get('total_flags', 0)}  "
        f"({summary.get('errors', 0)} errors, "
        f"{summary.get('warnings', 0)} warnings, "
        f"{summary.get('info_count', 0)} info)"
    )

    ws.merge_cells(f"B{row}:J{row}")
    sc = ws.cell(row, 2)
    sc.value     = summary_text
    sc.font      = Font(name="Arial", size=9)
    sc.fill      = PatternFill("solid", fgColor=ALT_ROW)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc.border    = thin_border()
    set_row_height(ws, row, 15)

    # Integrity score — shown prominently
    ws.merge_cells(f"K{row}:N{row}")
    ic = ws.cell(row, 11)
    ic.value     = f"  Data Integrity Score: {score}%"
    ic.font      = Font(name="Arial", bold=True, size=9, color=score_color)
    ic.fill      = PatternFill("solid", fgColor=score_bg)
    ic.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ic.border    = thin_border()
    row += 1

    # ── Flag table header ──────────────────────────────────────────────────────
    if not flags:
        # No flags — show green success row
        ws.merge_cells(f"B{row}:N{row}")
        nc = ws.cell(row, 2)
        nc.value     = "  ✓  No anomalies detected — data quality is excellent"
        nc.font      = Font(name="Arial", size=9, bold=True, color="1E6B1E")
        nc.fill      = PatternFill("solid", fgColor="E8F5E9")
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        nc.border    = thin_border()
        set_row_height(ws, row, 15)
        return

    # Column headers for flag table
    col_headers = ["Severity", "Method", "Metric", "Year", "Value", "Reason"]
    col_starts  = [2, 4, 6, 8, 10, 11]
    col_ends    = [3, 5, 7, 9, 10, 14]

    for header, cs, ce in zip(col_headers, col_starts, col_ends):
        if cs != ce:
            ws.merge_cells(f"{_col(cs)}{row}:{_col(ce)}{row}")
        c = ws.cell(row, cs)
        c.value     = f"  {header}"
        c.font      = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2C3E50")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = thin_border()
    set_row_height(ws, row, 14)
    row += 1

    # ── Flag rows ──────────────────────────────────────────────────────────────
    for fi, flag in enumerate(flags):
        sev      = flag.get("severity", "INFO")
        colors   = SEV_COLORS.get(sev, SEV_COLORS["INFO"])
        text_col = colors["text"]
        bg_col   = colors["bg"] if fi % 2 == 0 else "FFFFFF"

        value_str = f"₹{flag['value']:,.0f} Cr" if flag.get("value") is not None else "—"

        row_data = [
            (f"  {sev}",              col_starts[0], col_ends[0]),
            (f"  {flag.get('method','—')}",  col_starts[1], col_ends[1]),
            (f"  {flag.get('metric','—')}",  col_starts[2], col_ends[2]),
            (f"  {flag.get('year','—')}",    col_starts[3], col_ends[3]),
            (f"  {value_str}",               col_starts[4], col_ends[4]),
            (f"  {flag.get('reason','—')}",  col_starts[5], col_ends[5]),
        ]

        for val, cs, ce in row_data:
            if cs != ce:
                ws.merge_cells(f"{_col(cs)}{row}:{_col(ce)}{row}")
            c = ws.cell(row, cs)
            c.value     = val
            c.font      = Font(name="Arial", size=8, color=text_col,
                               bold=(val.strip() == sev))
            c.fill      = PatternFill("solid", fgColor=bg_col)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1, wrap_text=True)
            c.border    = thin_border()

        set_row_height(ws, row, 18)
        row += 1


def _col(n: int) -> str:
    """Convert column number to letter (1=A, 2=B ...)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _anomaly_rows(anomaly_report: dict) -> int:
    """Estimate how many rows the anomaly section occupies."""
    if not anomaly_report:
        return 0
    flags = anomaly_report.get("flags", [])
    # header + summary + col headers + flag rows + spacing
    return 4 + max(len(flags), 1) + 4


def _build_wacc_section(ws, wacc_result: dict, start_row: int):
    """
    Adds WACC COMPUTATION section to the Audit Trail sheet.
    Shows full CAPM breakdown so evaluators can see exactly how WACC was derived.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from excel.styles import thin_border, set_row_height, ALT_ROW

    WACC_COLOR = "1B5E20"   # dark green
    row = start_row + 2     # gap after anomaly section

    # ── Section header ─────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:N{row}")
    h = ws.cell(row, 2)
    h.value     = "WACC COMPUTATION  |  Dynamic CAPM-Based Discount Rate"
    h.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor=WACC_COLOR)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # ── WACC entries ───────────────────────────────────────────────────────────
    wacc_entries = [
        ("Beta",                 f"{wacc_result.get('beta', 0):.2f}  (source: {wacc_result.get('beta_source', '—')})"),
        ("Risk-Free Rate (Rf)",  f"{wacc_result.get('risk_free_rate', 0)*100:.1f}%  (10-yr Indian G-Sec)"),
        ("Equity Risk Premium",  f"{wacc_result.get('equity_risk_premium', 0)*100:.1f}%  (India standard)"),
        ("Cost of Equity (Ke)",  f"{wacc_result.get('cost_of_equity', 0)*100:.2f}%  [Rf + β × ERP]"),
        ("Cost of Debt (Kd)",    f"{wacc_result.get('cost_of_debt', 0)*100:.2f}%  [Interest / Borrowings]"),
        ("Tax Rate",             f"{wacc_result.get('tax_rate', 0)*100:.1f}%  [3-yr avg effective rate]"),
        ("Equity Weight (E/V)",  f"{wacc_result.get('equity_weight', 0)*100:.1f}%"),
        ("Debt Weight (D/V)",    f"{wacc_result.get('debt_weight', 0)*100:.1f}%"),
        ("Dynamic WACC",         f"{wacc_result.get('wacc', 0)*100:.2f}%  [E/V×Ke + D/V×Kd×(1-Tax)]"),
    ]

    for ri, (label, value) in enumerate(wacc_entries):
        alt  = (ri % 2 == 0)
        fill = PatternFill("solid", fgColor=ALT_ROW) if alt else None
        is_final = (label == "Dynamic WACC")

        lc = ws.cell(row, 2)
        vc = ws.cell(row, 5)

        lc.value     = f"  {label}"
        lc.font      = Font(name="Arial", size=9, bold=True,
                            color=WACC_COLOR if is_final else "000000")
        lc.border    = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"B{row}:D{row}")
        if fill: lc.fill = fill

        vc.value     = value
        vc.font      = Font(name="Arial", size=9, bold=is_final,
                            color=WACC_COLOR if is_final else "000000")
        vc.border    = thin_border()
        vc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"E{row}:N{row}")
        if fill: vc.fill = fill

        set_row_height(ws, row, 15)
        row += 1


def _clustering_rows(clustering_result: dict) -> int:
    """Estimate rows used by clustering section."""
    if not clustering_result:
        return 0
    return 14  # header + entries + spacing


def _build_clustering_section(ws, clustering_result: dict, start_row: int):
    """Adds INDUSTRY CLASSIFICATION section to the Audit Trail sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from excel.styles import thin_border, set_row_height, ALT_ROW

    CLUSTER_COLOR = "4A235A"   # purple
    row = start_row + 2

    ws.merge_cells(f"B{row}:N{row}")
    h = ws.cell(row, 2)
    h.value     = "INDUSTRY CLASSIFICATION  |  Distance-Based Sector Clustering"
    h.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor=CLUSTER_COLOR)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    entries = [
        ("Assigned Sector",      clustering_result.get("sector", "—")),
        ("Confidence",           f"{clustering_result.get('confidence', 0)*100:.0f}%"),
        ("Distance to Centroid", f"{clustering_result.get('distance', 0):.3f}"),
        ("Comparable Companies", ", ".join(clustering_result.get("peers", [])[:5])),
    ]

    # Distance to each sector
    for sector, dist in clustering_result.get("all_distances", {}).items():
        entries.append((f"  Distance → {sector}", f"{dist:.3f}"))

    for ri, (label, value) in enumerate(entries):
        alt  = (ri % 2 == 0)
        fill = PatternFill("solid", fgColor=ALT_ROW) if alt else None
        is_assigned = label == "Assigned Sector"

        lc = ws.cell(row, 2)
        vc = ws.cell(row, 5)

        lc.value     = f"  {label}"
        lc.font      = Font(name="Arial", size=9, bold=True,
                            color=CLUSTER_COLOR if is_assigned else "000000")
        lc.border    = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"B{row}:D{row}")
        if fill: lc.fill = fill

        vc.value     = value
        vc.font      = Font(name="Arial", size=9, bold=is_assigned,
                            color=CLUSTER_COLOR if is_assigned else "000000")
        vc.border    = thin_border()
        vc.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"E{row}:N{row}")
        if fill: vc.fill = fill

        set_row_height(ws, row, 15)
        row += 1