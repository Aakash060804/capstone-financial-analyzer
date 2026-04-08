"""
Writes AI Commentary sheet.
"""

from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from excel.styles       import (
    style_header, set_col_width, set_row_height,
    hide_gridlines, thin_border, NAVY, LIGHT_BLUE,
)
from ai.schemas import FullCommentaryReport
from utils.logger import get_logger

logger = get_logger(__name__)

TREND_COLORS = {
    "improving":    "1E6B1E",
    "stable":       "1F4E79",
    "deteriorating":"8B0000",
}


def build_commentary_sheet(wb: Workbook, report: FullCommentaryReport):
    ws = wb.create_sheet("AI Commentary")
    hide_gridlines(ws)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("B2:N2")
    style_header(ws.cell(2, 2),
                 f"AI-Generated Equity Research Commentary  |  {report.company}",
                 size=11)
    set_row_height(ws, 2, 22)

    ws.merge_cells("B3:N3")
    note = ws.cell(3, 2)
    note.value     = "⚠  Commentary is AI-generated based solely on computed ratios. Not investment advice."
    note.font      = Font(name="Arial", italic=True, size=8, color="8B0000")
    note.alignment = Alignment(horizontal="center")
    set_row_height(ws, 3, 14)

    row = 5

    # ── Investment Thesis ──────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:N{row}")
    t = ws.cell(row, 2)
    t.value     = "INVESTMENT THESIS"
    t.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # Rating
    rating_colors = {"Strong": "1E6B1E", "Adequate": "1F4E79", "Weak": "8B0000"}
    color = rating_colors.get(report.thesis.overall_rating, NAVY)
    ws.merge_cells(f"B{row}:N{row}")
    rc = ws.cell(row, 2)
    rc.value     = f"Overall Rating: {report.thesis.overall_rating}"
    rc.font      = Font(name="Arial", bold=True, size=11, color=color)
    rc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    set_row_height(ws, row, 20)
    row += 1

    # Executive summary
    ws.merge_cells(f"B{row}:N{row+3}")
    ec = ws.cell(row, 2)
    ec.value     = report.thesis.executive_summary
    ec.font      = Font(name="Arial", size=10)
    ec.alignment = Alignment(
        horizontal="left", vertical="top",
        wrap_text=True, indent=1
    )
    ec.border = thin_border()
    set_row_height(ws, row, 20)
    for r in range(row, row + 4):
        set_row_height(ws, r, 18)
    row += 5

    # Strengths and concerns side by side
    ws.cell(row, 2).value = "Key Strengths"
    ws.cell(row, 2).font  = Font(name="Arial", bold=True, size=9, color="1E6B1E")
    ws.cell(row, 8).value = "Key Concerns"
    ws.cell(row, 8).font  = Font(name="Arial", bold=True, size=9, color="8B0000")
    row += 1

    max_items = max(len(report.thesis.key_strengths), len(report.thesis.key_concerns))
    for i in range(max_items):
        if i < len(report.thesis.key_strengths):
            sc = ws.cell(row, 2)
            sc.value = f"  ✓  {report.thesis.key_strengths[i]}"
            sc.font  = Font(name="Arial", size=9, color="1E6B1E")
            ws.merge_cells(f"B{row}:G{row}")
        if i < len(report.thesis.key_concerns):
            cc = ws.cell(row, 8)
            cc.value = f"  ✗  {report.thesis.key_concerns[i]}"
            cc.font  = Font(name="Arial", size=9, color="8B0000")
            ws.merge_cells(f"H{row}:N{row}")
        set_row_height(ws, row, 15)
        row += 1

    row += 1

    # ── Category Commentaries ──────────────────────────────────────────────────
    if report.categories:
        ws.merge_cells(f"B{row}:N{row}")
        ch = ws.cell(row, 2)
        ch.value     = "CATEGORY ANALYSIS"
        ch.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ch.fill      = PatternFill("solid", fgColor=NAVY)
        ch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ch.border    = thin_border()
        set_row_height(ws, row, 18)
        row += 1

        for cat in report.categories:
            trend_color = TREND_COLORS.get(cat.trend.lower(), NAVY)

            # Category header
            ws.merge_cells(f"B{row}:N{row}")
            hc = ws.cell(row, 2)
            hc.value     = f"  {cat.category}  |  {cat.trend.upper()}"
            hc.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            hc.fill      = PatternFill("solid", fgColor=trend_color)
            hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            hc.border    = thin_border()
            set_row_height(ws, row, 16)
            row += 1

            # Headline
            ws.merge_cells(f"B{row}:N{row}")
            hl = ws.cell(row, 2)
            hl.value     = f"  {cat.headline}"
            hl.font      = Font(name="Arial", bold=True, size=9, color="1F4E79")
            hl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            set_row_height(ws, row, 15)
            row += 1

            # Commentary text
            ws.merge_cells(f"B{row}:N{row+1}")
            tc = ws.cell(row, 2)
            tc.value     = f"  {cat.commentary}"
            tc.font      = Font(name="Arial", size=9)
            tc.alignment = Alignment(
                horizontal="left", vertical="top",
                wrap_text=True, indent=1
            )
            tc.border = thin_border()
            set_row_height(ws, row, 18)
            set_row_height(ws, row + 1, 18)
            row += 3

    # ── Red Flags ──────────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"B{row}:N{row}")
    rf_header = ws.cell(row, 2)
    risk_colors = {"Low": "1E6B1E", "Medium": "C55A11", "High": "8B0000"}
    rf_color = risk_colors.get(report.red_flags.overall_risk, NAVY)
    rf_header.value     = f"RED FLAGS & RISK REGISTER  |  Overall Risk: {report.red_flags.overall_risk}"
    rf_header.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    rf_header.fill      = PatternFill("solid", fgColor=rf_color)
    rf_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rf_header.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    ws.merge_cells(f"B{row}:N{row}")
    rs = ws.cell(row, 2)
    rs.value     = f"  {report.red_flags.risk_summary}"
    rs.font      = Font(name="Arial", size=9)
    rs.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    rs.border    = thin_border()
    set_row_height(ws, row, 30)
    row += 2

    if report.red_flags.flags:
        sev_colors = {"high": "8B0000", "medium": "C55A11", "low": "1F4E79"}
        for flag in report.red_flags.flags:
            fc = ws.cell(row, 2)
            sc = sev_colors.get(flag.severity.lower(), NAVY)
            fc.value     = f"  [{flag.severity.upper()}]  {flag.metric}:  {flag.value}  —  {flag.explanation}"
            fc.font      = Font(name="Arial", size=9, color=sc)
            fc.alignment = Alignment(horizontal="left", wrap_text=True, indent=1)
            ws.merge_cells(f"B{row}:N{row}")
            set_row_height(ws, row, 20)
            row += 1
    else:
        nc = ws.cell(row, 2)
        nc.value = "  ✓  No red flags identified."
        nc.font  = Font(name="Arial", size=9, color="1E6B1E")
        ws.merge_cells(f"B{row}:N{row}")

    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 20)
    for ci in range(3, 15):
        set_col_width(ws, ci, 10)

    logger.info(f"Sheet 'AI Commentary' written")