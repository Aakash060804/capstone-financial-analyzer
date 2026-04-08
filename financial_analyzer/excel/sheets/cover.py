"""
Cover sheet — visual dashboard with key metrics scorecard.
"""

from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from openpyxl.utils    import get_column_letter
import pandas as pd
import numpy as np
from excel.styles       import (
    style_header, thin_border, set_col_width, set_row_height,
    hide_gridlines, NAVY, LIGHT_BLUE, ALT_ROW,
)
from config.settings import COMPANY_NAME
from utils.logger    import get_logger

logger = get_logger(__name__)

SCORECARD_METRICS = [
    ("Revenue Growth (%)",                  "Growth",       10.0,  True),
    ("EBITDA Margin (%)",                   "Profitability",15.0,  True),
    ("Net Profit Margin (%)",               "Profitability", 8.0,  True),
    ("Return on Equity % (ROE)",            "Returns",      15.0,  True),
    ("Return on Capital Employed % (ROCE)", "Returns",      12.0,  True),
    ("Current Ratio (x)",                   "Liquidity",    1.5,   True),
    ("Debt-to-Equity (x)",                  "Solvency",     1.0,   False),
    ("Interest Coverage (x)",               "Solvency",     5.0,   True),
    ("FCF Margin (%)",                      "Cash Flow",    5.0,   True),
]


def _traffic_light(value: float, threshold: float, higher_is_better: bool) -> str:
    """Returns hex color for traffic light."""
    if higher_is_better:
        if value >= threshold * 1.1:    return "1E6B1E"   # green
        if value >= threshold * 0.85:   return "C55A11"   # amber
        return "8B0000"                                     # red
    else:
        if value <= threshold * 0.9:    return "1E6B1E"
        if value <= threshold * 1.15:   return "C55A11"
        return "8B0000"


def build_cover_sheet(wb, ratio_df: pd.DataFrame, report, clustering_result: dict = None):
    ws = wb.active
    ws.title = "Cover"
    hide_gridlines(ws)

    # ── Main title block ───────────────────────────────────────────────────────
    ws.merge_cells("B2:O2")
    tc = ws.cell(2, 2)
    tc.value     = "FINANCIAL ANALYSIS REPORT"
    tc.font      = Font(name="Arial", bold=True, size=20, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 2, 40)

    ws.merge_cells("B3:O3")
    cc = ws.cell(3, 2)
    cc.value     = COMPANY_NAME
    cc.font      = Font(name="Arial", size=14, color=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 3, 25)

    ws.merge_cells("B4:O4")
    sub = ws.cell(4, 2)
    sub.value     = "Structured Financial Data  |  25+ Ratios  |  5-Factor DuPont  |  AI Commentary  |  DCF Valuation  |  Scenario Forecasts"
    sub.font      = Font(name="Arial", italic=True, size=9, color="595959")
    sub.alignment = Alignment(horizontal="center")
    set_row_height(ws, 4, 16)

    # ── Scorecard ──────────────────────────────────────────────────────────────
    mar_cols = [c for c in ratio_df.columns if "Mar" in str(c)]
    latest_yr = mar_cols[-1] if mar_cols else ""

    # ── Peer medians from clustering ─────────────────────────────────────────
    peer_medians = {}
    sector_label = ""
    if clustering_result:
        peer_medians = clustering_result.get("peer_medians", {})
        sector_label = clustering_result.get("sector", "")

    ws.merge_cells("B6:O6")
    sh = ws.cell(6, 2)
    sh.value     = f"KEY METRICS SCORECARD  |  {latest_yr}  |  Peer: {sector_label if sector_label else 'Generic'}"
    sh.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sh.fill      = PatternFill("solid", fgColor=NAVY)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sh.border    = thin_border()
    set_row_height(ws, 6, 18)

    # Scorecard headers
    headers = ["Metric", "Category", "Latest Value", "Threshold", "Signal"]
    widths  = [5, 2, 2, 2, 2]
    col_map = [2, 7, 10, 12, 14]
    merge_ends = [6, 9, 11, 13, 15]

    for i, h in enumerate(headers):
        ws.merge_cells(
            start_row=7, start_column=col_map[i],
            end_row=7, end_column=merge_ends[i]
        )
        style_header(ws.cell(7, col_map[i]), h, size=9)
    set_row_height(ws, 7, 16)

    for ri, (metric, category, threshold, higher) in enumerate(SCORECARD_METRICS):
        row_num = 8 + ri
        alt     = (ri % 2 == 0)
        fill    = PatternFill("solid", fgColor=ALT_ROW) if alt else None

        val = None
        if metric in ratio_df.index:
            # Get last non-null value across all Mar years
            mar_cols = [c for c in ratio_df.columns if "Mar" in str(c)]
            for yr in reversed(mar_cols):
                if yr in ratio_df.columns:
                    v = ratio_df.loc[metric, yr]
                    if not pd.isna(v):
                        val = float(v)
                        break

        # Metric name
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=6)
        mc = ws.cell(row_num, 2)
        mc.value     = f"  {metric}"
        mc.font      = Font(name="Arial", size=9)
        mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        mc.border    = thin_border()
        if fill: mc.fill = fill

        # Category
        ws.merge_cells(start_row=row_num, start_column=7, end_row=row_num, end_column=9)
        catc = ws.cell(row_num, 7)
        catc.value     = category
        catc.font      = Font(name="Arial", size=9, color="595959")
        catc.alignment = Alignment(horizontal="center")
        catc.border    = thin_border()
        if fill: catc.fill = fill

        # Value
        ws.merge_cells(start_row=row_num, start_column=10, end_row=row_num, end_column=11)
        vc = ws.cell(row_num, 10)
        if val is not None:
            vc.value          = val / 100 if "%" in metric else val
            vc.number_format  = '0.0%' if "%" in metric else '0.00x'
        else:
            vc.value = "N/A"
        vc.font      = Font(name="Arial", size=9, bold=True)
        vc.alignment = Alignment(horizontal="center")
        vc.border    = thin_border()
        if fill: vc.fill = fill

        # Threshold — peer median if available, else generic
        effective_threshold = peer_medians.get(metric, threshold)
        is_peer_calibrated  = metric in peer_medians
        ws.merge_cells(start_row=row_num, start_column=12, end_row=row_num, end_column=13)
        tc2 = ws.cell(row_num, 12)
        if "%" in metric:
            thresh_label = f"{effective_threshold*100:.1f}%"
        else:
            thresh_label = f"{effective_threshold:.2f}x"
        tc2.value     = thresh_label + (" (peer)" if is_peer_calibrated else " (generic)")
        tc2.font      = Font(name="Arial", size=8, color="1F4E79" if is_peer_calibrated else "595959")
        tc2.alignment = Alignment(horizontal="center")
        tc2.border    = thin_border()
        if fill: tc2.fill = fill

        # Signal — peer-calibrated traffic light
        ws.merge_cells(start_row=row_num, start_column=14, end_row=row_num, end_column=15)
        sig = ws.cell(row_num, 14)
        if val is not None:
            color  = _traffic_light(val, effective_threshold, higher)
            symbol = "●  STRONG" if color == "1E6B1E" else ("●  WATCH" if color == "C55A11" else "●  WEAK")
        else:
            color, symbol = "595959", "N/A"
        sig.value     = symbol
        sig.font      = Font(name="Arial", size=9, bold=True, color=color)
        sig.alignment = Alignment(horizontal="center")
        sig.border    = thin_border()
        if fill: sig.fill = fill

        set_row_height(ws, row_num, 15)

    # ── Sheet index ────────────────────────────────────────────────────────────
    index_row = 8 + len(SCORECARD_METRICS) + 2

    ws.merge_cells(f"B{index_row}:O{index_row}")
    ih = ws.cell(index_row, 2)
    ih.value     = "WORKBOOK INDEX"
    ih.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ih.fill      = PatternFill("solid", fgColor=NAVY)
    ih.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ih.border    = thin_border()
    set_row_height(ws, index_row, 18)
    index_row += 1

    sheets_info = [
        ("Income Statement",  "Profit & Loss — Revenue, EBITDA, PAT (vertical layout)"),
        ("Balance Sheet",     "Assets, Liabilities, Equity — vertical layout"),
        ("Cash Flow",         "Operating, Investing, Financing Cash Flows"),
        ("Financial Ratios",  "25+ ratios across 6 categories with CAGR column"),
        ("DuPont Analysis",   "5-factor ROE decomposition with variance check"),
        ("AI Commentary",     "Equity research commentary + red flag register"),
        ("Forecasts",         "Base/Bull/Bear scenarios + DCF + Sensitivity table"),
    ]

    for i, (sheet, desc) in enumerate(sheets_info):
        alt  = (i % 2 == 0)
        fill = PatternFill("solid", fgColor=ALT_ROW) if alt else None

        ws.merge_cells(f"B{index_row}:F{index_row}")
        sc = ws.cell(index_row, 2)
        sc.value     = f"  {sheet}"
        sc.font      = Font(name="Arial", bold=True, size=9, color=NAVY)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc.border    = thin_border()
        if fill: sc.fill = fill

        ws.merge_cells(f"G{index_row}:O{index_row}")
        dc = ws.cell(index_row, 7)
        dc.value     = desc
        dc.font      = Font(name="Arial", size=9)
        dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dc.border    = thin_border()
        if fill: dc.fill = fill

        set_row_height(ws, index_row, 15)
        index_row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    set_col_width(ws, 1, 2)
    for ci in range(2, 16):
        set_col_width(ws, ci, 9)
    set_col_width(ws, 2, 3)

    logger.info("Sheet 'Cover' written")