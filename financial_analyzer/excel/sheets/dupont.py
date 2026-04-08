"""
Writes DuPont Analysis sheet.
"""

import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from excel.styles       import (
    style_header, style_label, style_value,
    set_col_width, set_row_height, freeze, hide_gridlines,
    thin_border, NAVY, LIGHT_BLUE, ALT_ROW,
)
from utils.logger import get_logger

logger = get_logger(__name__)

# Format: (type, is_bold, is_variance, already_whole_number)
# already_whole_number=True means value is like 18.3, needs /100 for Excel %
# already_whole_number=False means value is like 0.054, just apply % format directly
FACTOR_FORMATS = {
    "(1) Tax Burden  [NI / EBT]":                           ("ratio", False, False, False),
    "(2) Interest Burden  [EBT / EBIT]":                    ("ratio", False, False, False),
    "(3) EBIT Margin  [EBIT / Revenue]":                    ("pct",   False, False, False),
    "(4) Asset Turnover  [Revenue / Avg Assets]":           ("ratio", False, False, False),
    "(5) Equity Multiplier  [Avg Assets / Avg Equity]":     ("ratio", False, False, False),
    "DuPont ROE (%)  [Product of above × 100]":             ("pct",   True,  False, True),
    "Direct ROE (%)  [NI / Avg Equity × 100]":              ("pct",   True,  False, True),
    "Variance (DuPont vs Direct)":                          ("pct",   False, True,  True),
}


def build_dupont_sheet(wb: Workbook, dupont_df: pd.DataFrame):
    ws = wb.create_sheet("DuPont Analysis")
    hide_gridlines(ws)
    freeze(ws, "C5")

    mar_cols = [c for c in dupont_df.columns if "Mar" in str(c)]
    dupont_df = dupont_df[mar_cols]
    n_years   = len(mar_cols)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + n_years)
    style_header(
        ws.cell(2, 2),
        "5-Factor DuPont Analysis  |  ROE = Tax Burden × Interest Burden × EBIT Margin × Asset Turnover × Equity Multiplier",
        size=10
    )
    set_row_height(ws, 2, 22)

    # ── Column headers ─────────────────────────────────────────────────────────
    ws.cell(4, 2).value     = "DuPont Factor"
    ws.cell(4, 2).font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(4, 2).fill      = PatternFill("solid", fgColor=NAVY)
    ws.cell(4, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(4, 2).border    = thin_border()
    for ci, yr in enumerate(mar_cols):
        style_header(ws.cell(4, 3 + ci), str(yr), size=9)
    set_row_height(ws, 4, 18)

    # ── Data rows ──────────────────────────────────────────────────────────────
    for ri, factor in enumerate(dupont_df.index):
        row_num = 5 + ri
        alt     = (ri % 2 == 0)
        fmt = FACTOR_FORMATS.get(factor, ("ratio", False, False, False))
        fmt_type, is_bold, is_variance, is_whole = fmt

        lc = ws.cell(row_num, 2)
        style_label(lc, f"  {factor}", bold=is_bold, alt=alt)
        if is_bold:
            lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

        for ci, yr in enumerate(mar_cols):
            vc  = ws.cell(row_num, 3 + ci)
            val = dupont_df.loc[factor, yr]
            v   = float(val) if not pd.isna(val) else None
            
            is_pct   = fmt_type == "pct"
            is_ratio = fmt_type == "ratio"

            if is_pct and v is not None:
                if is_whole:
                    v = v / 100    # stored as 18.3 → send 0.183 to Excel % format
                # else: stored as 0.054 → send as-is to Excel % format
                
            style_value(vc, value=v, is_pct=is_pct, is_ratio=is_ratio, alt=alt)

            if is_bold and v is not None:
                vc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                vc.font = Font(name="Arial", size=9, bold=True)

            if is_variance and v is not None:
                # Color variance: green if near zero, red if large
                color = "1E6B1E" if abs(float(val)) < 0.01 else "8B0000"
                vc.font = Font(name="Arial", size=9, color=color)

        set_row_height(ws, row_num, 15)

    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 48)
    for ci in range(n_years):
        set_col_width(ws, 3 + ci, 12)

    logger.info(f"Sheet 'DuPont Analysis' written: {len(dupont_df)} factors")