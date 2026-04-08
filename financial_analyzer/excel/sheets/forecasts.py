"""
Writes Forecasts sheet — Scenarios + DCF + Sensitivity Table.
"""

import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from excel.styles       import (
    style_header, style_label, style_value, style_assumption,
    set_col_width, set_row_height, hide_gridlines,
    thin_border, col_letter, NAVY, ALT_ROW,
    SCENARIO_COLORS,
)
from utils.logger import get_logger

logger = get_logger(__name__)

PCT_METRICS = {
    "EBITDA Margin (%)", "EBIT Margin (%)",
    "Net Margin (%)", "FCF Margin (%)",
    "ROE (%) est.", "Revenue Growth (%)",
}

RAW_PCT_METRICS = {"OPM %", "Tax %"}


def build_forecasts_sheet(
    wb: Workbook,
    scenarios: dict,
    dcf_result: dict,
    sensitivity_df: pd.DataFrame,
    prophet_result: dict = None,
):
    ws = wb.create_sheet("Forecasts")
    hide_gridlines(ws)

    row = 2

    # ── Master title ───────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:R{row}")
    style_header(ws.cell(row, 2),
                "Financial Forecasts  |  Scenarios  |  DCF Valuation  |  Sensitivity Analysis",
                size=11)
    set_row_height(ws, row, 22)
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1: Scenario Assumptions
    # ══════════════════════════════════════════════════════════════════════════
    from config.settings import SCENARIO_ASSUMPTIONS, FORECAST_YEARS

    ws.merge_cells(f"B{row}:H{row}")
    sh = ws.cell(row, 2)
    sh.value     = "SCENARIO ASSUMPTIONS"
    sh.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sh.fill      = PatternFill("solid", fgColor=NAVY)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sh.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    headers = ["Scenario", "Revenue Growth", "EBIT Margin Δ", "Tax Rate", "Forecast Years"]
    for ci, h in enumerate(headers):
        style_header(ws.cell(row, 2 + ci), h, size=9)
    set_row_height(ws, row, 16)
    row += 1

    for scenario, a in SCENARIO_ASSUMPTIONS.items():
        vals = [
            scenario.capitalize(),
            a["revenue_growth"],
            a["ebit_margin_delta"],
            a["tax_rate"],
            FORECAST_YEARS,
        ]
        for ci, v in enumerate(vals):
            c = ws.cell(row, 2 + ci)
            if ci == 0:
                c.value     = v
                c.font      = Font(name="Arial", bold=True, size=9,
                                    color="FFFFFF")
                c.fill      = PatternFill("solid",
                                        fgColor=SCENARIO_COLORS.get(scenario, NAVY))
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin_border()
            else:
                style_assumption(c, v if ci == 4 else None)
                if ci in [1, 2, 3]:
                    c.value          = v
                    c.number_format  = '0.0%'
                else:
                    c.value = v
        set_row_height(ws, row, 15)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2: Scenario P&L projections
    # ══════════════════════════════════════════════════════════════════════════
    for scenario, df in scenarios.items():
        color   = SCENARIO_COLORS.get(scenario, NAVY)
        n_years = len(df.columns)

        ws.merge_cells(f"B{row}:{col_letter(2+n_years)}{row}")
        sc = ws.cell(row, 2)
        sc.value     = f"{scenario.upper()} CASE  |  ₹ Crores"
        sc.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        sc.fill      = PatternFill("solid", fgColor=color)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc.border    = thin_border()
        set_row_height(ws, row, 18)
        row += 1

        # Year headers
        ws.cell(row, 2).value     = "Metric"
        ws.cell(row, 2).font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws.cell(row, 2).fill      = PatternFill("solid", fgColor=color)
        ws.cell(row, 2).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row, 2).border    = thin_border()

        for ci, yr in enumerate(df.columns):
            style_header(ws.cell(row, 3 + ci), str(yr), bg=color, size=9)
        set_row_height(ws, row, 16)
        row += 1

        for ri, metric in enumerate(df.index):
            alt    = (ri % 2 == 0)
            is_pct = metric in PCT_METRICS

            lc = ws.cell(row, 2)
            style_label(lc, f"  {metric}", alt=alt)

            for ci, yr in enumerate(df.columns):
                vc  = ws.cell(row, 3 + ci)
                val = df.loc[metric, yr]
                v   = float(val) if not pd.isna(val) else None
                if is_pct and v is not None:
                    v = v / 100
                    
                is_raw_pct = metric in RAW_PCT_METRICS
                style_value(vc, value=v, is_pct=is_pct,
                            is_int=not is_pct and not is_raw_pct, alt=alt)
                if is_raw_pct and vc.value is not None:
                    vc.number_format = '0.0"%"'

            set_row_height(ws, row, 15)
            row += 1

        row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3: DCF Valuation
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"B{row}:J{row}")
    dh = ws.cell(row, 2)
    dh.value     = "DCF VALUATION SUMMARY"
    dh.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    dh.fill      = PatternFill("solid", fgColor="4A235A")
    dh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    dh.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # Assumptions block
    for k, v in dcf_result["assumptions"].items():
        lc = ws.cell(row, 2)
        vc = ws.cell(row, 4)
        lc.value     = f"  {k}"
        lc.font      = Font(name="Arial", size=9)
        lc.alignment = Alignment(horizontal="left", indent=1)
        lc.border    = thin_border()
        ws.merge_cells(f"B{row}:C{row}")

        style_assumption(vc, v)
        ws.merge_cells(f"D{row}:E{row}")
        set_row_height(ws, row, 15)
        row += 1

    row += 1

    # Projected FCF table
    summary_df = dcf_result["summary_df"]
    style_header(ws.cell(row, 2), "Year", size=9)
    style_header(ws.cell(row, 3), "Projected FCF (₹ Cr)", size=9)
    style_header(ws.cell(row, 4), "PV of FCF (₹ Cr)", size=9)
    set_row_height(ws, row, 16)
    row += 1

    for ri, yr in enumerate(summary_df.index):
        alt = (ri % 2 == 0)
        ws.cell(row, 2).value = yr
        ws.cell(row, 2).font  = Font(name="Arial", size=9)
        ws.cell(row, 2).border = thin_border()
        if alt:
            ws.cell(row, 2).fill = PatternFill("solid", fgColor=ALT_ROW)

        for ci, col in enumerate(summary_df.columns):
            vc = ws.cell(row, 3 + ci)
            style_value(vc, value=float(summary_df.loc[yr, col]),
                        is_int=True, alt=alt)
        set_row_height(ws, row, 15)
        row += 1

    # Key outputs
    row += 1
    key_outputs = [
        ("Terminal Value (₹ Cr)",      dcf_result["terminal_value"]),
        ("PV of Terminal Value (₹ Cr)",dcf_result["pv_terminal_value"]),
        ("Enterprise Value (₹ Cr)",    dcf_result["enterprise_value"]),
        ("Net Debt (₹ Cr)",            dcf_result["net_debt"]),
        ("Equity Value (₹ Cr)",        dcf_result["equity_value"]),
        ("Intrinsic Value / Share (₹)",dcf_result["intrinsic_value_per_share"]),
    ]
    for label, val in key_outputs:
        lc = ws.cell(row, 2)
        vc = ws.cell(row, 4)
        lc.value     = f"  {label}"
        lc.font      = Font(name="Arial", size=9, bold="Intrinsic" in label)
        lc.border    = thin_border()
        ws.merge_cells(f"B{row}:C{row}")
        style_value(vc, value=float(val), is_int=True)
        if "Intrinsic" in label:
            vc.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        ws.merge_cells(f"D{row}:E{row}")
        set_row_height(ws, row, 15)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3b: Monte Carlo Simulation Results  ← AI FEATURE 2
    # ══════════════════════════════════════════════════════════════════════════
    mc = dcf_result.get("monte_carlo")
    if mc:
        ws.merge_cells(f"B{row}:J{row}")
        mc_header = ws.cell(row, 2)
        mc_header.value     = f"MONTE CARLO SIMULATION  |  {mc['n_simulations']:,} Runs  |  Intrinsic Value Distribution"
        mc_header.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        mc_header.fill      = PatternFill("solid", fgColor="1F3864")
        mc_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        mc_header.border    = thin_border()
        set_row_height(ws, row, 18)
        row += 1

        # Explanation note
        ws.merge_cells(f"B{row}:J{row}")
        note = ws.cell(row, 2)
        note.value = (
            f"  WACC sampled from Normal({mc['wacc_used']:.1%}, ±1.5%)  |  "
            f"TGR sampled from Normal({mc['tgr_used']:.1%}, ±0.8%)  |  "
            f"Both bounded to realistic ranges.  "
            f"Probability of fair value within ±20% of median: {mc['prob_within_20']:.0f}%"
        )
        note.font      = Font(name="Arial", italic=True, size=8, color="595959")
        note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        set_row_height(ws, row, 14)
        row += 1

        row += 1

        # Three-column probability display
        mc_rows = [
            ("Pessimistic",  "25th Percentile",  mc["label_p25"],  "8B0000", "FDECEA",
             f"WACC toward upper range (~{mc['wacc_used']*100+1.5:.0f}%+)"),
            ("Most Likely",  "Median (50th Pct)", mc["label_p50"],  "1F4E79", "EBF5FB",
             f"Base case WACC ~{mc['wacc_used']*100:.0f}%"),
            ("Optimistic",   "75th Percentile",   mc["label_p75"],  "1E6B1E", "E8F5E9",
             f"WACC toward lower range (~{max(6, mc['wacc_used']*100-1.5):.0f}%-)"),
        ]

        col_starts = [2, 6, 10]
        col_widths = [4, 4, 4]  # each scenario spans 4 columns

        for (scenario, pct_label, value, text_color, bg_color, note_text), col_start in zip(mc_rows, col_starts):
            col_end = col_start + col_widths[0] - 1
            col_s   = col_letter(col_start)
            col_e   = col_letter(col_end)

            # Scenario label header
            ws.merge_cells(f"{col_s}{row}:{col_e}{row}")
            sh = ws.cell(row, col_start)
            sh.value     = scenario
            sh.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            sh.fill      = PatternFill("solid", fgColor=text_color)
            sh.alignment = Alignment(horizontal="center", vertical="center")
            sh.border    = thin_border()

        set_row_height(ws, row, 20)
        row += 1

        for (scenario, pct_label, value, text_color, bg_color, note_text), col_start in zip(mc_rows, col_starts):
            col_end = col_start + col_widths[0] - 1
            col_s   = col_letter(col_start)
            col_e   = col_letter(col_end)

            # Percentile label
            ws.merge_cells(f"{col_s}{row}:{col_e}{row}")
            pl = ws.cell(row, col_start)
            pl.value     = pct_label
            pl.font      = Font(name="Arial", size=8, color="595959", italic=True)
            pl.fill      = PatternFill("solid", fgColor=bg_color)
            pl.alignment = Alignment(horizontal="center", vertical="center")
            pl.border    = thin_border()

        set_row_height(ws, row, 14)
        row += 1

        for (scenario, pct_label, value, text_color, bg_color, note_text), col_start in zip(mc_rows, col_starts):
            col_end = col_start + col_widths[0] - 1
            col_s   = col_letter(col_start)
            col_e   = col_letter(col_end)

            # Big value
            ws.merge_cells(f"{col_s}{row}:{col_e}{row}")
            vc = ws.cell(row, col_start)
            vc.value     = value
            vc.font      = Font(name="Arial", bold=True, size=16, color=text_color)
            vc.fill      = PatternFill("solid", fgColor=bg_color)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border    = thin_border()

        set_row_height(ws, row, 28)
        row += 1

        for (scenario, pct_label, value, text_color, bg_color, note_text), col_start in zip(mc_rows, col_starts):
            col_end = col_start + col_widths[0] - 1
            col_s   = col_letter(col_start)
            col_e   = col_letter(col_end)

            # Note
            ws.merge_cells(f"{col_s}{row}:{col_e}{row}")
            nc = ws.cell(row, col_start)
            nc.value     = note_text
            nc.font      = Font(name="Arial", size=8, color="595959", italic=True)
            nc.fill      = PatternFill("solid", fgColor=bg_color)
            nc.alignment = Alignment(horizontal="center", vertical="center")
            nc.border    = thin_border()

        set_row_height(ws, row, 14)
        row += 2

        # Summary stats row
        ws.merge_cells(f"B{row}:J{row}")
        stats = ws.cell(row, 2)
        stats.value = (
            f"  Mean: {mc['label_p50']}  |  "
            f"Std Dev: ₹{mc['std']:,.0f}  |  "
            f"10th pct: {mc['label_p10']}  |  "
            f"90th pct: {mc['label_p90']}  |  "
            f"Valid simulations: {mc['n_simulations']:,}"
        )
        stats.font      = Font(name="Arial", size=8, color="595959")
        stats.alignment = Alignment(horizontal="left", indent=1)
        stats.border    = thin_border()
        set_row_height(ws, row, 14)
        row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4: Sensitivity Table
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"B{row}:L{row}")
    sh2 = ws.cell(row, 2)
    sh2.value     = "SENSITIVITY ANALYSIS  |  Intrinsic Value per Share (₹)  |  WACC vs Terminal Growth Rate"
    sh2.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sh2.fill      = PatternFill("solid", fgColor="154360")
    sh2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sh2.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # Header row
    ws.cell(row, 2).value     = "WACC \\ TGR →"
    ws.cell(row, 2).font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(row, 2).fill      = PatternFill("solid", fgColor="154360")
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    ws.cell(row, 2).border    = thin_border()

    for ci, col in enumerate(sensitivity_df.columns):
        style_header(ws.cell(row, 3 + ci), str(col), bg="154360", size=9)
    set_row_height(ws, row, 16)
    row += 1

    # Find base case value for conditional coloring
    base_val = dcf_result["intrinsic_value_per_share"]

    for ri, wacc_label in enumerate(sensitivity_df.index):
        alt = (ri % 2 == 0)
        lc  = ws.cell(row, 2)
        style_label(lc, str(wacc_label), bold=True, alt=alt)

        for ci, col in enumerate(sensitivity_df.columns):
            vc  = ws.cell(row, 3 + ci)
            val = sensitivity_df.loc[wacc_label, col]

            if pd.isna(val):
                vc.value = "N/A"
                vc.font  = Font(name="Arial", size=9, color="595959")
            else:
                v = float(val)
                style_value(vc, value=v, is_int=True, alt=alt)
                # Color code relative to base case
                if v >= base_val * 1.1:
                    vc.font = Font(name="Arial", size=9, color="1E6B1E")
                elif v <= base_val * 0.9:
                    vc.font = Font(name="Arial", size=9, color="8B0000")

            vc.border = thin_border()

        set_row_height(ws, row, 15)
        row += 1

    # Column widths
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 28)
    for ci in range(10):
        set_col_width(ws, 3 + ci, 14)

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 5 — Prophet Time Series Forecast + Comparison
    # ═══════════════════════════════════════════════════════════════════════
    if prophet_result and prophet_result.get("available"):
        row += 2
        row = _build_prophet_section(ws, row, prophet_result, scenarios)

    logger.info("Sheet 'Forecasts' written")

# ─────────────────────────────────────────────────────────────────────────────
# PROPHET SECTION BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_prophet_section(ws, start_row: int, prophet_result: dict, scenarios: dict) -> int:
    """
    Adds two blocks to the Forecasts sheet after the sensitivity table:

    Block A — Prophet Forecast Table
        Revenue / EBITDA / FCF with actual history + forecast + lower/upper bounds
        Forecast years shaded blue. Actual years in grey.

    Block B — Comparison Table
        Assumptions-based Base/Bull/Bear vs Prophet Base/Bull/Bear
        Side by side with difference row.

    Returns the next available row number.
    """
    PROPHET_COLOR = "0D47A1"   # deep blue header
    COMPARE_COLOR = "1B5E20"   # deep green header
    ACTUAL_BG     = "F5F5F5"   # light grey for actual years
    FORECAST_BG   = "EBF5FB"   # light blue for forecast years

    row = start_row

    # ── BLOCK A HEADER ─────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:R{row}")
    h = ws.cell(row, 2)
    h.value     = (
        "FACEBOOK PROPHET  |  Time Series Forecast  |  "
        "Revenue · EBITDA · Free Cash Flow  |  80% Confidence Interval"
    )
    h.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor=PROPHET_COLOR)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # Note row
    ws.merge_cells(f"B{row}:R{row}")
    note = ws.cell(row, 2)
    note.value     = (
        "  Prophet trained on 14 years of actual data.  "
        "Upper bound = Bull (80th pct).  Lower bound = Bear (20th pct).  "
        "Shaded blue = forecast years.  All values in ₹ Crores."
    )
    note.font      = Font(name="Arial", italic=True, size=8, color="595959")
    note.alignment = Alignment(horizontal="left", indent=1)
    set_row_height(ws, row, 13)
    row += 1

    # ── FOR EACH METRIC ────────────────────────────────────────────────────────
    metric_map = {
        "revenue": "Revenue (₹ Cr)",
        "ebitda":  "EBITDA (₹ Cr)",
        "fcf":     "Free Cash Flow (₹ Cr)",
    }

    for key, label in metric_map.items():
        if key not in prophet_result:
            continue

        df = prophet_result[key]
        if df.empty:
            continue

        all_years = df["year"].tolist()
        n_cols    = len(all_years)

        # Metric sub-header
        ws.merge_cells(f"B{row}:{col_letter(2 + n_cols)}{row}")
        mh = ws.cell(row, 2)
        mh.value     = f"  {label}"
        mh.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        mh.fill      = PatternFill("solid", fgColor=PROPHET_COLOR)
        mh.alignment = Alignment(horizontal="left", indent=1)
        mh.border    = thin_border()
        set_row_height(ws, row, 15)
        row += 1

        # Year header row
        ws.cell(row, 2).value      = "Series"
        ws.cell(row, 2).font       = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        ws.cell(row, 2).fill       = PatternFill("solid", fgColor=PROPHET_COLOR)
        ws.cell(row, 2).border     = thin_border()
        ws.cell(row, 2).alignment  = Alignment(horizontal="left", indent=1)

        for ci, yr in enumerate(all_years):
            style_header(ws.cell(row, 3 + ci), str(yr), bg=PROPHET_COLOR, size=8)
        set_row_height(ws, row, 14)
        row += 1

        # Three data rows: Forecast, Lower, Upper
        series_rows = [
            ("Forecast (Base)", "value", "1F4E79"),
            ("Lower — Bear",    "lower", "8B0000"),
            ("Upper — Bull",    "upper", "1E6B1E"),
        ]

        for series_label, col_key, text_color in series_rows:
            lc = ws.cell(row, 2)
            lc.value     = f"  {series_label}"
            lc.font      = Font(name="Arial", size=8, bold=True, color=text_color)
            lc.border    = thin_border()
            lc.alignment = Alignment(horizontal="left", indent=1)

            for ci, yr in enumerate(all_years):
                vc       = ws.cell(row, 3 + ci)
                row_data = df[df["year"] == yr]

                if not row_data.empty:
                    is_forecast = row_data["type"].iloc[0] == "forecast"
                    val         = float(row_data[col_key].iloc[0])
                    vc.value         = round(val, 0)
                    vc.number_format = "#,##0"
                    vc.font          = Font(
                        name="Arial", size=8,
                        color=text_color if is_forecast else "595959",
                        bold=is_forecast,
                    )
                    vc.fill = PatternFill("solid", fgColor=FORECAST_BG if is_forecast else ACTUAL_BG)
                else:
                    vc.value = "—"
                    vc.font  = Font(name="Arial", size=8, color="AAAAAA")

                vc.alignment = Alignment(horizontal="right")
                vc.border    = thin_border()

            set_row_height(ws, row, 14)
            row += 1

        row += 1

    # ── BLOCK B — COMPARISON TABLE ─────────────────────────────────────────────
    row += 1

    ws.merge_cells(f"B{row}:R{row}")
    ch = ws.cell(row, 2)
    ch.value     = (
        "FORECAST COMPARISON  |  "
        "Assumptions-Based  vs  Prophet Time Series  |  Revenue (₹ Cr)"
    )
    ch.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ch.fill      = PatternFill("solid", fgColor=COMPARE_COLOR)
    ch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ch.border    = thin_border()
    set_row_height(ws, row, 18)
    row += 1

    # Get forecast years from base scenario
    if "base" not in scenarios:
        return row

    base_df    = scenarios["base"]
    fore_years = base_df.columns.tolist()
    n_fy       = len(fore_years)

    # Column headers
    ws.cell(row, 2).value     = "  Method"
    ws.cell(row, 2).font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(row, 2).fill      = PatternFill("solid", fgColor=COMPARE_COLOR)
    ws.cell(row, 2).border    = thin_border()
    ws.cell(row, 2).alignment = Alignment(horizontal="left", indent=1)

    for ci, yr in enumerate(fore_years):
        style_header(ws.cell(row, 3 + ci), str(yr), bg=COMPARE_COLOR, size=9)
    set_row_height(ws, row, 15)
    row += 1

    def write_row(label, values, text_color, bg_color, bold=False):
        nonlocal row
        lc = ws.cell(row, 2)
        lc.value     = f"  {label}"
        lc.font      = Font(name="Arial", size=9, bold=bold, color=text_color)
        lc.fill      = PatternFill("solid", fgColor=bg_color)
        lc.border    = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)

        for ci, v in enumerate(values):
            vc = ws.cell(row, 3 + ci)
            if v is not None:
                vc.value         = round(v, 0)
                vc.number_format = "#,##0"
            else:
                vc.value = "—"
            vc.font      = Font(name="Arial", size=9, color=text_color, bold=bold)
            vc.fill      = PatternFill("solid", fgColor=bg_color)
            vc.alignment = Alignment(horizontal="right")
            vc.border    = thin_border()
        set_row_height(ws, row, 15)
        row += 1

    # Assumptions rows
    for scenario, text_c, bg_c in [
        ("base", "1F4E79", "EBF5FB"),
        ("bull", "1E6B1E", "E8F5E9"),
        ("bear", "8B0000", "FDECEA"),
    ]:
        if scenario in scenarios:
            sdf  = scenarios[scenario]
            vals = [
                sdf.loc["Revenue", yr] if yr in sdf.columns else None
                for yr in fore_years
            ]
            write_row(f"Assumptions — {scenario.capitalize()}", vals, text_c, bg_c)

    # Separator
    for ci in range(n_fy + 1):
        ws.cell(row, 2 + ci).value = ""
    set_row_height(ws, row, 5)
    row += 1

    # Prophet rows
    if "revenue" in prophet_result:
        pdf          = prophet_result["revenue"]
        forecast_pdf = pdf[pdf["type"] == "forecast"]

        for col_key, label, text_c, bg_c in [
            ("value", "Prophet — Base", "1F4E79", "EBF5FB"),
            ("upper", "Prophet — Bull", "1E6B1E", "E8F5E9"),
            ("lower", "Prophet — Bear", "8B0000", "FDECEA"),
        ]:
            vals = []
            for yr in fore_years:
                r = forecast_pdf[forecast_pdf["year"] == yr]
                vals.append(float(r[col_key].iloc[0]) if not r.empty else None)
            write_row(label, vals, text_c, bg_c, bold=True)

    # Difference row
    if "revenue" in prophet_result and "base" in scenarios:
        pdf          = prophet_result["revenue"]
        forecast_pdf = pdf[pdf["type"] == "forecast"]

        lc = ws.cell(row, 2)
        lc.value     = "  Difference  (Prophet Base − Assumptions Base)"
        lc.font      = Font(name="Arial", size=8, italic=True, color="595959")
        lc.border    = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)

        for ci, yr in enumerate(fore_years):
            vc    = ws.cell(row, 3 + ci)
            p_row = forecast_pdf[forecast_pdf["year"] == yr]
            p_val = float(p_row["value"].iloc[0]) if not p_row.empty else None
            a_val = base_df.loc["Revenue", yr] if yr in base_df.columns else None

            if p_val is not None and a_val is not None:
                diff             = p_val - a_val
                vc.value         = round(diff, 0)
                vc.number_format = "#,##0"
                vc.font          = Font(
                    name="Arial", size=8, italic=True,
                    color="1E6B1E" if diff >= 0 else "8B0000"
                )
            else:
                vc.value = "—"
                vc.font  = Font(name="Arial", size=8, color="AAAAAA")

            vc.alignment = Alignment(horizontal="right")
            vc.border    = thin_border()

        set_row_height(ws, row, 13)
        row += 1

    return row