"""
Writes Financial Ratios sheet — vertical layout with category groupings.
"""

import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles   import Font, PatternFill, Alignment
from excel.styles       import (
    style_header, style_label, style_value, style_section_label,
    set_col_width, set_row_height, freeze, hide_gridlines,
    thin_border, col_letter, NAVY, LIGHT_BLUE, ALT_ROW,
)
from utils.logger import get_logger

logger = get_logger(__name__)

CATEGORIES = {
    "PROFITABILITY": [
        "EBITDA Margin (%)",
        "EBIT Margin (%)",
        "Net Profit Margin (%)",
        "Return on Assets % (ROA)",
        "Return on Equity % (ROE)",
        "Return on Capital Employed % (ROCE)",
    ],
    "ASSET UTILIZATION": [
        "Asset Turnover (x)",
        "Fixed Asset Turnover (x)",
        "Inventory Turnover (x)",
        "Receivables Turnover (x)",
        "Days Inventory Outstanding (days)",
        "Days Sales Outstanding (days)",
        "Days Payable Outstanding (days)",
        "Cash Conversion Cycle (days)",
    ],
    "LIQUIDITY": [
        "Current Ratio (x)",
        "Quick Ratio (x)",
        "Cash Ratio (x)",
    ],
    "SOLVENCY": [
        "Debt-to-Equity (x)",
        "Debt-to-Assets (x)",
        "Interest Coverage (x)",
        "Net Debt (₹ Cr)",
        "Net Debt / EBITDA (x)",
    ],
    "CASH FLOW": [
        "Operating CF Margin (%)",
        "Free Cash Flow (₹ Cr)",
        "FCF Margin (%)",
        "FCF to Net Income (x)",
        "CapEx Intensity (%)",
    ],
    "GROWTH": [
        "Revenue Growth (%)",
        "EBITDA Growth (%)",
        "Net Income Growth (%)",
        "EPS Growth (%)",
        "CFO Growth (%)",
    ],
}

CATEGORY_COLORS = {
    "PROFITABILITY":     "1F4E79",
    "ASSET UTILIZATION": "375623",
    "LIQUIDITY":         "7B3F00",
    "SOLVENCY":          "4A235A",
    "CASH FLOW":         "154360",
    "GROWTH":            "145A32",
}


def _classify_format(ratio_name: str):
    is_pct   = "%" in ratio_name
    is_ratio = "(x)" in ratio_name
    is_days  = "days" in ratio_name.lower() or "Days" in ratio_name
    is_int   = "₹" in ratio_name and not is_pct
    return is_pct, is_ratio, is_days, is_int


def build_ratios_sheet(wb: Workbook, ratio_df: pd.DataFrame, clustering_result: dict = None):
    ws = wb.create_sheet("Financial Ratios")
    hide_gridlines(ws)
    freeze(ws, "C5")

    mar_cols = [c for c in ratio_df.columns if "Mar" in str(c)]
    ratio_df = ratio_df[mar_cols]
    n_years  = len(mar_cols)

    # ── Peer medians from clustering ──────────────────────────────────────────
    peer_medians = {}
    if clustering_result:
        peer_medians = clustering_result.get("peer_medians", {})

    # ── CAGR column ────────────────────────────────────────────────────────────
    # Compute CAGR for each ratio over full period
    def compute_cagr(series: pd.Series) -> float:
        vals = series.dropna()
        if len(vals) < 2:
            return np.nan
        start, end = float(vals.iloc[0]), float(vals.iloc[-1])
        n = len(vals) - 1
        if start <= 0 or end <= 0:
            return np.nan
        return ((end / start) ** (1 / n) - 1) * 100

    # ── Title ──────────────────────────────────────────────────────────────────
    has_peers  = bool(peer_medians)
    total_cols = n_years + 1 + (1 if has_peers else 0)  # years + CAGR + optional peer median
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + total_cols + 1)
    style_header(ws.cell(2, 2), f"Financial Ratios  |  {ratio_df.columns[-1]} Latest", size=11)
    set_row_height(ws, 2, 22)

    # ── Column headers ─────────────────────────────────────────────────────────
    ws.cell(4, 2).value     = "Ratio"
    ws.cell(4, 2).font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(4, 2).fill      = PatternFill("solid", fgColor=NAVY)
    ws.cell(4, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(4, 2).border    = thin_border()

    for ci, yr in enumerate(mar_cols):
        style_header(ws.cell(4, 3 + ci), str(yr), size=9)

    cagr_header = ws.cell(4, 3 + n_years)
    style_header(cagr_header, "CAGR", size=9, bg="2E4057")

    if has_peers:
        peer_header = ws.cell(4, 4 + n_years)
        style_header(peer_header, "Industry Median", size=9, bg="1B5E20")

    set_row_height(ws, 4, 18)

    # ── Data rows by category ──────────────────────────────────────────────────
    row_num = 5

    for category, ratio_names in CATEGORIES.items():
        color = CATEGORY_COLORS.get(category, NAVY)

        # Category separator row
        cat_cell = ws.cell(row_num, 2)
        style_section_label(cat_cell, f"  {category}", color=color)
        for ci in range(total_cols):
            c = ws.cell(row_num, 3 + ci)
            c.fill   = PatternFill("solid", fgColor=color)
            c.border = thin_border()
        set_row_height(ws, row_num, 16)
        row_num += 1

        cat_ri = 0
        for ratio_name in ratio_names:
            if ratio_name not in ratio_df.index:
                continue

            alt = (cat_ri % 2 == 0)
            is_pct, is_ratio, is_days, is_int = _classify_format(ratio_name)

            # Label
            lc = ws.cell(row_num, 2)
            style_label(lc, f"  {ratio_name}", alt=alt)

            # Values
            for ci, yr in enumerate(mar_cols):
                vc  = ws.cell(row_num, 3 + ci)
                val = ratio_df.loc[ratio_name, yr]
                v   = float(val) if not pd.isna(val) else None

                # Convert pct values to decimal for Excel percentage format
                if is_pct and v is not None:
                    v = v / 100

                style_value(
                    vc, value=v,
                    is_pct=is_pct,
                    is_ratio=is_ratio,
                    is_days=is_days,
                    is_int=is_int,
                    alt=alt,
                )

            # CAGR
            cagr_val  = compute_cagr(ratio_df.loc[ratio_name])
            cagr_cell = ws.cell(row_num, 3 + n_years)
            if cagr_val is not None and not np.isnan(cagr_val):
                cagr_cell.value = cagr_val / 100
                cagr_cell.number_format = '0.0%;(0.0%);"-"'
            else:
                cagr_cell.value = None
                cagr_cell.number_format = '"-"'
            cagr_cell.font      = Font(name="Arial", size=9, color="2E4057")
            cagr_cell.alignment = Alignment(horizontal="right")
            cagr_cell.border    = thin_border()
            if alt:
                cagr_cell.fill = PatternFill("solid", fgColor=ALT_ROW)

            # Industry Median column
            if has_peers:
                peer_cell = ws.cell(row_num, 4 + n_years)
                peer_val  = peer_medians.get(ratio_name)
                if peer_val is not None:
                    if is_pct:
                        peer_cell.value         = peer_val
                        peer_cell.number_format = '0.0%'
                    elif is_ratio:
                        peer_cell.value         = peer_val
                        peer_cell.number_format = '0.00x'
                    else:
                        peer_cell.value         = peer_val
                        peer_cell.number_format = '#,##0'
                    peer_cell.font      = Font(name="Arial", size=9, color="1B5E20", bold=True)
                else:
                    peer_cell.value = "—"
                    peer_cell.font  = Font(name="Arial", size=9, color="AAAAAA")
                peer_cell.alignment = Alignment(horizontal="right")
                peer_cell.border    = thin_border()
                if alt:
                    peer_cell.fill = PatternFill("solid", fgColor="E8F5E9")

            set_row_height(ws, row_num, 15)
            row_num += 1
            cat_ri  += 1

        row_num += 1  # blank gap between categories

    # ── Column widths ──────────────────────────────────────────────────────────
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 38)
    for ci in range(n_years):
        set_col_width(ws, 3 + ci, 12)
    set_col_width(ws, 3 + n_years, 10)  # CAGR column
    if has_peers:
        set_col_width(ws, 4 + n_years, 16)  # Industry Median column

    logger.info(f"Sheet 'Financial Ratios' written: {row_num - 5} rows")