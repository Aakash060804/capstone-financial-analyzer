"""
Writes Income Statement, Balance Sheet, and Cash Flow
as vertical tables: rows = line items, columns = years.
"""

import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles   import PatternFill, Font, Alignment
from excel.styles       import (
    style_header, style_label, style_value, style_section_label,
    set_col_width, set_row_height, freeze, hide_gridlines,
    thin_border, col_letter, NAVY, LIGHT_BLUE, ALT_ROW, WHITE
)
from utils.logger import get_logger

logger = get_logger(__name__)

# Screener's canonical row order for each statement
# Only rows in this list will appear, in this exact order
INCOME_ROW_ORDER = [
    "Sales",
    "Expenses",
    "Operating Profit",
    "OPM %",
    "Other Income",
    "Interest",
    "Depreciation",
    "Profit before tax",
    "Tax %",
    "Net Profit",
    "EPS in Rs",
    "Dividend Payout %",
]

BALANCE_ROW_ORDER = [
    "Equity Capital",
    "Reserves",
    "Borrowings",
    "Other Liabilities",
    "Total Liabilities",
    "Fixed Assets",
    "CWIP",
    "Investments",
    "Other Assets",
    "Total Assets",
]

CASHFLOW_ROW_ORDER = [
    "Cash from Operating Activity",
    "Profit from operations",
    "Receivables",
    "Inventory",
    "Payables",
    "Loans Advances",
    "Other WC items",
    "Working capital changes",
    "Direct taxes",
    "Cash from Investing Activity",
    "Fixed assets purchased",
    "Fixed assets sold",
    "Investments purchased",
    "Investments sold",
    "Interest received",
    "Dividends received",
    "Acquisition of companies",
    "Other investing items",
    "Cash from Financing Activity",
    "Proceeds from borrowings",
    "Repayment of borrowings",
    "Interest paid fin",
    "Dividends paid",
    "Financial liabilities",
    "Other financing items",
    "Net Cash Flow",
]

BOLD_ROWS_MAP = {
    "Income Statement": [
        "Sales", "Operating Profit", "Profit before tax", "Net Profit"
    ],
    "Balance Sheet": [
        "Total Liabilities", "Total Assets", "Borrowings"
    ],
    "Cash Flow": [
        "Cash from Operating Activity",
        "Working capital changes",
        "Cash from Investing Activity",
        "Cash from Financing Activity",
        "Net Cash Flow",
    ],
}


def _write_statement(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    title: str,
    unit_note: str,
    row_order: list[str],
):
    ws = wb.create_sheet(sheet_name)
    hide_gridlines(ws)
    freeze(ws, "C5")

    # Filter to Mar years only
    mar_cols = [c for c in df.columns if "Mar" in str(c)]
    df = df[mar_cols]
    n_years = len(mar_cols)

    # Build ordered row list — only rows that exist in data
    ordered_rows = []
    for row_label in row_order:
        if row_label in df.index:
            ordered_rows.append(row_label)
            
    bold_rows = BOLD_ROWS_MAP.get(sheet_name, [])

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells(
        start_row=2, start_column=2,
        end_row=2, end_column=2 + n_years
    )
    style_header(ws.cell(2, 2), f"{title}  |  {unit_note}", size=11)
    set_row_height(ws, 2, 22)

    ws.merge_cells(
        start_row=3, start_column=2,
        end_row=3, end_column=2 + n_years
    )
    sub = ws.cell(3, 2)
    sub.value     = "All figures in ₹ Crores unless stated otherwise"
    sub.font      = Font(name="Arial", italic=True, size=8, color="595959")
    sub.alignment = Alignment(horizontal="center")
    set_row_height(ws, 3, 14)

    # ── Year headers ───────────────────────────────────────────────────────────
    import openpyxl
    lh = ws.cell(4, 2)
    lh.value     = "Metric"
    lh.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    lh.fill      = openpyxl.styles.PatternFill("solid", fgColor=NAVY)
    lh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lh.border    = thin_border()

    for ci, yr in enumerate(mar_cols):
        style_header(ws.cell(4, 3 + ci), str(yr), size=9)
    set_row_height(ws, 4, 18)

    # ── Data rows ──────────────────────────────────────────────────────────────
    RAW_PCT_ROWS = {"Tax %", "OPM %", "Dividend Payout %"}

    for ri, metric in enumerate(ordered_rows):
        row_num  = 5 + ri
        alt      = (ri % 2 == 0)
        is_bold  = metric in bold_rows
        is_raw_pct = metric in RAW_PCT_ROWS
        import openpyxl as _ox

        lc = ws.cell(row_num, 2)
        style_label(lc, str(metric), bold=is_bold, indent=1, alt=alt)
        if is_bold:
            lc.fill = _ox.styles.PatternFill("solid", fgColor=LIGHT_BLUE)

        for ci, yr in enumerate(mar_cols):
            vc  = ws.cell(row_num, 3 + ci)
            val = df.loc[metric, yr]

            if pd.isna(val):
                vc.value = None
            else:
                vc.value = float(val)

            style_value(vc, is_int=not is_raw_pct, alt=alt)

            if is_raw_pct and vc.value is not None:
                vc.number_format = '0.0"%"'

            if is_bold and not pd.isna(val):
                vc.fill = _ox.styles.PatternFill("solid", fgColor=LIGHT_BLUE)
                vc.font = Font(name="Arial", size=9, bold=True)

        set_row_height(ws, row_num, 15)

    # ── Column widths ──────────────────────────────────────────────────────────
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 32)
    for ci in range(n_years):
        set_col_width(ws, 3 + ci, 12)

    logger.info(f"Sheet '{sheet_name}' written: {len(ordered_rows)} rows x {n_years} years")


def build_income_statement(wb: Workbook, income_df: pd.DataFrame):
    _write_statement(
        wb, income_df,
        sheet_name="Income Statement",
        title="Profit & Loss Statement",
        unit_note="₹ Crores",
        row_order=INCOME_ROW_ORDER,
    )

def build_balance_sheet(wb: Workbook, balance_df: pd.DataFrame):
    _write_statement(
        wb, balance_df,
        sheet_name="Balance Sheet",
        title="Balance Sheet",
        unit_note="₹ Crores",
        row_order=BALANCE_ROW_ORDER,
    )

def build_cash_flow(wb: Workbook, cashflow_df: pd.DataFrame):
    _write_statement(
        wb, cashflow_df,
        sheet_name="Cash Flow",
        title="Cash Flow Statement",
        unit_note="₹ Crores",
        row_order=CASHFLOW_ROW_ORDER,
    )