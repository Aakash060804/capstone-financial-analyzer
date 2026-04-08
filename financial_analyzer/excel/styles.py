"""
All colors, fonts, borders, and number formats.
Single source of truth — imported by all sheet builders.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter


# ── Color palette ──────────────────────────────────────────────────────────────
NAVY        = "1F4E79"
WHITE       = "FFFFFF"
LIGHT_BLUE  = "D6E4F0"
ALT_ROW     = "F0F7FF"
BLUE_INPUT  = "0000FF"
BLACK       = "000000"
GREEN_LINK  = "375623"
YELLOW_KEY  = "FFFF00"
RED_FLAG    = "FF0000"
GREY_TEXT   = "595959"
DARK_GREEN  = "1E6B1E"
DARK_RED    = "8B0000"
ORANGE      = "C55A11"

SCENARIO_COLORS = {
    "base": "1F4E79",
    "bull": "375623",
    "bear": "843C0C",
}


# ── Border helpers ─────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def thick_bottom_border():
    thin = Side(style="thin",   color="BFBFBF")
    thck = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=thck)


# ── Cell style appliers ────────────────────────────────────────────────────────

def style_header(cell, text=None, bg=NAVY, fg=WHITE, size=10, bold=True, center=True):
    if text is not None:
        cell.value = text
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True
    )
    cell.border = thin_border()


def style_label(cell, text=None, bold=False, indent=0, alt=False):
    if text is not None:
        cell.value = text
    cell.font      = Font(name="Arial", size=9, bold=bold, color=BLACK)
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=indent
    )
    cell.border = thin_border()
    if alt:
        cell.fill = PatternFill("solid", fgColor=ALT_ROW)


def style_value(cell, value=None, is_pct=False, is_ratio=False,
                is_int=False, is_days=False, alt=False, color=BLACK):
    if value is not None:
        cell.value = value
    cell.font      = Font(name="Arial", size=9, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = thin_border()
    if alt:
        cell.fill  = PatternFill("solid", fgColor=ALT_ROW)

    if is_pct:
        cell.number_format = '0.0%;(0.0%);"-"'
    elif is_ratio:
        cell.number_format = '0.00x;(0.00x);"-"'
    elif is_days:
        cell.number_format = '0.0;(0.0);"-"'
    elif is_int:
        cell.number_format = '#,##0;(#,##0);"-"'
    else:
        cell.number_format = '#,##0.0;(#,##0.0);"-"'


def style_section_label(cell, text, color=NAVY):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border    = thin_border()


def style_assumption(cell, value=None):
    if value is not None:
        cell.value = value
    cell.font      = Font(name="Arial", size=9, color=BLUE_INPUT)
    cell.fill      = PatternFill("solid", fgColor=YELLOW_KEY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()


def col_letter(col_idx: int) -> str:
    return get_column_letter(col_idx)


def set_col_width(ws, col_idx: int, width: float):
    ws.column_dimensions[col_letter(col_idx)].width = width


def set_row_height(ws, row_idx: int, height: float):
    ws.row_dimensions[row_idx].height = height


def freeze(ws, cell="C4"):
    ws.freeze_panes = cell


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False